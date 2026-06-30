#!/usr/bin/env python3
"""
nav_eval.py — Navigation2 goal-arrival accuracy evaluator (ROS 2 Jazzy).

For each predefined goal pose, send a NavigateToPose action to Nav2, capture
the actual arrival pose from ground-truth (TF or a pose topic), compute
translation/heading errors, and log every trial to:
  - a CSV file (appended live, crash-safe)
  - an XLSX file (written at end of run, with a formatted 'trials' sheet
    and a 'summary' sheet of per-goal stats computed via Excel formulas)

Pass criteria (configurable):
  e_trans  <= 0.05 m   (±5 cm)
  |e_head| <= 5.0 deg  (±5°)

Usage:
  python3 nav_eval.py --config config.yaml
"""

import argparse
import csv
import math
import sys
from datetime import datetime
from pathlib import Path

import rclpy
import yaml
from geometry_msgs.msg import PoseWithCovarianceStamped, Quaternion
from nav2_msgs.action import NavigateToPose
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from rclpy.action import ActionClient
from rclpy.duration import Duration
from rclpy.node import Node
from tf2_ros import Buffer, TransformException, TransformListener


HEADER = [
    'timestamp', 'trial_id', 'goal_id',
    'x_goal', 'y_goal', 'theta_goal_deg',
    'x_actual', 'y_actual', 'theta_actual_deg',
    'e_trans_m', 'e_heading_deg', 'pass',
    'nav_status', 'notes',
]


# ---------- geometry helpers ----------

def yaw_from_quat(q: Quaternion) -> float:
    siny_cosp = 2.0 * (q.w * q.z + q.x * q.y)
    cosy_cosp = 1.0 - 2.0 * (q.y * q.y + q.z * q.z)
    return math.atan2(siny_cosp, cosy_cosp)


def quat_from_yaw(yaw: float) -> Quaternion:
    q = Quaternion()
    q.z = math.sin(yaw / 2.0)
    q.w = math.cos(yaw / 2.0)
    return q


def wrap_angle(a: float) -> float:
    return math.atan2(math.sin(a), math.cos(a))


# ---------- ROS 2 node ----------

class NavEvaluator(Node):
    def __init__(self, cfg: dict):
        super().__init__('nav_eval')
        self.cfg = cfg
        self.map_frame = cfg.get('map_frame', 'map')
        self.base_frame = cfg.get('base_frame', 'base_link')
        self.pose_source = cfg.get('pose_source', 'tf')        # 'tf' or 'topic'
        self.pose_topic = cfg.get('pose_topic', '/amcl_pose')
        self.last_pose = None

        self.tf_buffer = Buffer()
        self.tf_listener = TransformListener(self.tf_buffer, self)

        if self.pose_source == 'topic':
            self.create_subscription(
                PoseWithCovarianceStamped, self.pose_topic, self._pose_cb, 10
            )
            self.get_logger().info(f'Ground-truth pose from topic: {self.pose_topic}')
        else:
            self.get_logger().info(
                f'Ground-truth pose from TF: {self.map_frame} -> {self.base_frame}'
            )

        self.nav_client = ActionClient(self, NavigateToPose, 'apple/navigate_to_pose')

    def _pose_cb(self, msg: PoseWithCovarianceStamped):
        self.last_pose = msg.pose.pose

    def wait_for_nav(self, timeout_sec: float = 30.0) -> bool:
        self.get_logger().info('Waiting for Nav2 action server...')
        return self.nav_client.wait_for_server(timeout_sec=timeout_sec)

    def get_actual_pose(self, settle_sec: float = 1.0):
        end_time = self.get_clock().now() + Duration(seconds=settle_sec)
        while rclpy.ok() and self.get_clock().now() < end_time:
            rclpy.spin_once(self, timeout_sec=0.05)

        if self.pose_source == 'tf':
            try:
                tr = self.tf_buffer.lookup_transform(
                    self.map_frame, self.base_frame, rclpy.time.Time()
                )
            except TransformException as e:
                self.get_logger().error(f'TF lookup failed: {e}')
                return None
            t = tr.transform.translation
            q = tr.transform.rotation
            return (t.x, t.y, yaw_from_quat(q))

        if self.last_pose is None:
            self.get_logger().error('No pose received on topic yet.')
            return None
        p = self.last_pose
        return (p.position.x, p.position.y, yaw_from_quat(p.orientation))

    def send_goal(self, x: float, y: float, theta: float) -> str:
        goal_msg = NavigateToPose.Goal()
        goal_msg.pose.header.frame_id = self.map_frame
        goal_msg.pose.header.stamp = self.get_clock().now().to_msg()
        goal_msg.pose.pose.position.x = float(x)
        goal_msg.pose.pose.position.y = float(y)
        goal_msg.pose.pose.orientation = quat_from_yaw(float(theta))

        send_future = self.nav_client.send_goal_async(goal_msg)
        rclpy.spin_until_future_complete(self, send_future)
        goal_handle = send_future.result()
        if goal_handle is None or not goal_handle.accepted:
            return 'REJECTED'

        result_future = goal_handle.get_result_async()
        rclpy.spin_until_future_complete(self, result_future)
        result = result_future.result()
        status = result.status if result is not None else 0
        # action_msgs/GoalStatus: 4=SUCCEEDED, 5=CANCELED, 6=ABORTED
        return {4: 'SUCCEEDED', 5: 'CANCELED', 6: 'ABORTED'}.get(status, f'STATUS_{status}')


# ---------- CSV streaming logger (crash safety) ----------

class CsvWriter:
    def __init__(self, csv_path: Path | None):
        self.path = csv_path
        self._fp = None
        self._writer = None
        if csv_path is None:
            return
        csv_path.parent.mkdir(parents=True, exist_ok=True)
        new_file = not csv_path.exists()
        self._fp = open(csv_path, 'a', newline='')
        self._writer = csv.writer(self._fp)
        if new_file:
            self._writer.writerow(HEADER)
            self._fp.flush()

    def write(self, row: list):
        if self._writer is None:
            return
        out = ['TRUE' if v is True else 'FALSE' if v is False else v for v in row]
        self._writer.writerow(out)
        self._fp.flush()

    def close(self):
        if self._fp is not None:
            self._fp.close()


# ---------- XLSX writer ----------

def write_xlsx(path: Path, rows: list, goal_ids: list) -> None:
    """Write a formatted workbook with 'trials' and 'summary' sheets."""
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # ===== Sheet 1: trials =====
    ws = wb.active
    ws.title = 'trials'
    ws.append(HEADER)

    head_font = Font(name='Arial', bold=True, color='FFFFFF')
    head_fill = PatternFill('solid', start_color='305496')
    head_align = Alignment(horizontal='center', vertical='center')
    for col_idx in range(1, len(HEADER) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = head_font
        c.fill = head_fill
        c.alignment = head_align

    body_font = Font(name='Arial')
    for row in rows:
        ws.append(row)
    for r in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for c in r:
            c.font = body_font

    widths = {
        'A': 20, 'B': 9, 'C': 10,
        'D': 10, 'E': 10, 'F': 14,
        'G': 10, 'H': 10, 'I': 14,
        'J': 12, 'K': 14, 'L': 8,
        'M': 12, 'N': 14,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = 'A2'

    # Conditional formatting on pass column (L): green TRUE / red FALSE
    n = len(rows)
    if n > 0:
        rng = f'L2:L{n + 1}'
        green = PatternFill('solid', start_color='C6EFCE')
        red = PatternFill('solid', start_color='FFC7CE')
        ws.conditional_formatting.add(rng, FormulaRule(formula=['L2=TRUE'], fill=green))
        ws.conditional_formatting.add(rng, FormulaRule(formula=['L2=FALSE'], fill=red))

    # Number formats on numeric cells
    for r in range(2, n + 2):
        for col in ('D', 'E', 'G', 'H'):
            ws[f'{col}{r}'].number_format = '0.000'
        for col in ('F', 'I'):
            ws[f'{col}{r}'].number_format = '0.00'
        ws[f'J{r}'].number_format = '0.0000'
        ws[f'K{r}'].number_format = '0.000'

    # ===== Sheet 2: summary =====
    s = wb.create_sheet('summary')
    last = n + 1  # last data row in trials sheet

    title_font = Font(name='Arial', bold=True, size=12)
    sub_font = Font(name='Arial', bold=True)
    sub_fill = PatternFill('solid', start_color='D9E1F2')

    s['A1'] = 'Navigation accuracy summary'
    s['A1'].font = title_font

    if n == 0:
        s['A3'] = 'No trial data.'
        wb.save(path)
        return

    cR = f'trials!$C$2:$C${last}'   # goal_id
    jR = f'trials!$J$2:$J${last}'   # e_trans_m
    kR = f'trials!$K$2:$K${last}'   # e_heading_deg
    lR = f'trials!$L$2:$L${last}'   # pass

    # --- Overall block ---
    s['A3'] = 'Overall'
    s['A3'].font = sub_font
    s['A3'].fill = sub_fill

    overall_hdr = [
        'Total trials', 'Passed', 'Pass rate',
        'Mean e_trans (m)', 'Max e_trans (m)',
        'Mean |e_head| (deg)', 'Max |e_head| (deg)',
    ]
    for i, h in enumerate(overall_hdr, start=1):
        c = s.cell(row=4, column=i, value=h)
        c.font = sub_font
        c.fill = sub_fill
        c.alignment = head_align

    s.cell(row=5, column=1, value=f'=COUNTA({cR})')
    s.cell(row=5, column=2, value=f'=COUNTIF({lR}, TRUE)')
    s.cell(row=5, column=3, value='=IFERROR(B5/A5, 0)')
    s.cell(row=5, column=3).number_format = '0.0%'
    s.cell(row=5, column=4, value=f'=IFERROR(AVERAGE({jR}), 0)')
    s.cell(row=5, column=4).number_format = '0.0000'
    s.cell(row=5, column=5, value=f'=IFERROR(MAX({jR}), 0)')
    s.cell(row=5, column=5).number_format = '0.0000'
    s.cell(row=5, column=6, value=f'=IFERROR(SUMPRODUCT(ABS({kR}))/A5, 0)')
    s.cell(row=5, column=6).number_format = '0.000'
    s.cell(row=5, column=7, value=f'=IFERROR(SUMPRODUCT(MAX(ABS({kR}))), 0)')
    s.cell(row=5, column=7).number_format = '0.000'

    # --- Per-goal block ---
    s['A7'] = 'Per goal'
    s['A7'].font = sub_font
    s['A7'].fill = sub_fill

    pg_hdr = [
        'goal_id', 'trials', 'passed', 'pass rate',
        'mean e_trans (m)', 'max e_trans (m)',
        'mean |e_head| (deg)', 'max |e_head| (deg)',
    ]
    for i, h in enumerate(pg_hdr, start=1):
        c = s.cell(row=8, column=i, value=h)
        c.font = sub_font
        c.fill = sub_fill
        c.alignment = head_align

    row_idx = 9
    for gid in goal_ids:
        s.cell(row=row_idx, column=1, value=gid)
        s.cell(row=row_idx, column=2, value=f'=COUNTIF({cR}, "{gid}")')
        s.cell(row=row_idx, column=3, value=f'=COUNTIFS({cR}, "{gid}", {lR}, TRUE)')
        s.cell(row=row_idx, column=4, value=f'=IFERROR(C{row_idx}/B{row_idx}, 0)')
        s.cell(row=row_idx, column=4).number_format = '0.0%'
        s.cell(row=row_idx, column=5, value=f'=IFERROR(AVERAGEIF({cR}, "{gid}", {jR}), 0)')
        s.cell(row=row_idx, column=5).number_format = '0.0000'
        s.cell(row=row_idx, column=6,
               value=f'=IFERROR(SUMPRODUCT(MAX(({cR}="{gid}")*{jR})), 0)')
        s.cell(row=row_idx, column=6).number_format = '0.0000'
        s.cell(row=row_idx, column=7,
               value=f'=IFERROR(SUMPRODUCT(({cR}="{gid}")*ABS({kR}))/B{row_idx}, 0)')
        s.cell(row=row_idx, column=7).number_format = '0.000'
        s.cell(row=row_idx, column=8,
               value=f'=IFERROR(SUMPRODUCT(MAX(({cR}="{gid}")*ABS({kR}))), 0)')
        s.cell(row=row_idx, column=8).number_format = '0.000'
        for col in range(1, 9):
            s.cell(row=row_idx, column=col).font = body_font
        row_idx += 1

    for col, w in [('A', 14), ('B', 10), ('C', 10), ('D', 12),
                   ('E', 18), ('F', 18), ('G', 22), ('H', 22)]:
        s.column_dimensions[col].width = w
    s.freeze_panes = 'A9'

    wb.save(path)


# ---------- main ----------

def main():
    parser = argparse.ArgumentParser(description='Nav2 goal-arrival evaluator.')
    parser.add_argument('--config', '-c', required=True, help='/home/rgt/test_navigation/config.example.yaml')
    args = parser.parse_args()

    with open(args.config, 'r') as f:
        cfg = yaml.safe_load(f)

    e_trans_thresh = float(cfg.get('e_trans_threshold_m', 0.05))
    e_head_thresh = float(cfg.get('e_heading_threshold_deg', 5.0))
    repeats = int(cfg.get('repeats_per_goal', 3))
    settle = float(cfg.get('settle_seconds', 1.0))
    interactive = bool(cfg.get('interactive', False))
    xlsx_path = Path(cfg.get('xlsx_path', 'nav_eval_log.xlsx'))
    csv_path_cfg = cfg.get('csv_path')
    csv_path = Path(csv_path_cfg) if csv_path_cfg else None

    goals = cfg.get('goals') or []
    if not goals:
        print('No goals defined in config.', file=sys.stderr)
        sys.exit(1)
    goal_ids = [str(g['id']) for g in goals]

    rclpy.init()
    node = NavEvaluator(cfg)
    if not node.wait_for_nav(timeout_sec=30.0):
        node.get_logger().error('Nav2 action server not available; exiting.')
        node.destroy_node()
        rclpy.shutdown()
        sys.exit(2)

    csv_writer = CsvWriter(csv_path)
    rows = []

    trial_id = 0
    total = 0
    passed = 0
    xlsx_msg = ''

    try:
        for g in goals:
            gid = str(g['id'])
            xg = float(g['x'])
            yg = float(g['y'])
            tg = float(g['theta'])
            tg_deg = math.degrees(tg)

            for k in range(repeats):
                trial_id += 1
                total += 1

                if interactive:
                    input(
                        f'\n[Trial {trial_id}] Reset robot to start, then press '
                        f'ENTER to send goal "{gid}" (run {k + 1}/{repeats})...'
                    )

                node.get_logger().info(
                    f'Trial {trial_id}: goal {gid} -> '
                    f'({xg:.3f} m, {yg:.3f} m, {tg_deg:.1f} deg) [{k + 1}/{repeats}]'
                )

                status = node.send_goal(xg, yg, tg)
                actual = node.get_actual_pose(settle_sec=settle)
                ts = datetime.now().isoformat(timespec='seconds')

                if actual is None:
                    row = [
                        ts, trial_id, gid,
                        round(xg, 4), round(yg, 4), round(tg_deg, 3),
                        '', '', '',
                        '', '', False,
                        status, 'no_pose',
                    ]
                else:
                    xa, ya, ta = actual
                    e_t = math.hypot(xa - xg, ya - yg)
                    e_h_deg = math.degrees(wrap_angle(ta - tg))
                    ok = (
                        status == 'SUCCEEDED'
                        and e_t <= e_trans_thresh
                        and abs(e_h_deg) <= e_head_thresh
                    )
                    if ok:
                        passed += 1
                    row = [
                        ts, trial_id, gid,
                        round(xg, 4), round(yg, 4), round(tg_deg, 3),
                        round(xa, 4), round(ya, 4), round(math.degrees(ta), 3),
                        round(e_t, 4), round(e_h_deg, 3),
                        bool(ok),
                        status, '',
                    ]

                rows.append(row)
                csv_writer.write(row)
                node.get_logger().info(
                    f'  result: pass={row[11]}  e_trans={row[9]} m  '
                    f'e_head={row[10]} deg  status={status}'
                )

    except KeyboardInterrupt:
        node.get_logger().info('Interrupted by user; writing what we have...')
    finally:
        csv_writer.close()
        try:
            write_xlsx(xlsx_path, rows, goal_ids)
            xlsx_msg = f'XLSX log   : {xlsx_path.resolve()}'
        except Exception as e:
            xlsx_msg = f'XLSX write FAILED: {e}'
        node.destroy_node()
        rclpy.shutdown()

    rate = (100.0 * passed / total) if total else 0.0
    print('\n========== SUMMARY ==========')
    print(f'Trials run : {total}')
    print(f'Passed     : {passed}')
    print(f'Pass rate  : {rate:.1f}%')
    if csv_path is not None:
        print(f'CSV log    : {csv_path.resolve()}')
    print(xlsx_msg)


if __name__ == '__main__':
    main()
